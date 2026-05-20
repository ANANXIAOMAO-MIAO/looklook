"""
LOOKLOOK — 文献抓取与智能速递
"""

from __future__ import annotations

import html
import json
import re
import time
from datetime import date, timedelta
from io import BytesIO
from typing import Any

import requests
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

DEEPSEEK_API_URL = "https://api.deepseek.com/chat/completions"
DEEPSEEK_MODEL = "deepseek-chat"
OPENALEX_SEARCH_URL = "https://api.openalex.org/works"
OPENALEX_PER_PAGE = 200
OPENALEX_SLEEP = 0.1
MAX_FETCH_PAPERS = 100
MAX_SELECTED_PAPERS = 10
MAX_RETRIES = 3
RETRY_INTERVAL = 2

TIME_RANGE_OPTIONS = ["一月内", "半年内", "一年内", "三年内", "五年内"]
DEFAULT_TIME_RANGE = "三年内"

TIME_RANGE_DAYS = {
    "一月内": 30,
    "半年内": 182,
    "一年内": 365,
    "三年内": 1095,
    "五年内": 1825,
}

SUBTITLE = "基于 OpenAlex + DeepSeek，用五分钟和感兴趣的前沿文献打个照面"

# ---------------------------------------------------------------------------
# 页面配置
# ---------------------------------------------------------------------------

st.set_page_config(
    page_title="LOOKLOOK",
    page_icon="📚",
    layout="wide",
    initial_sidebar_state="collapsed",
)

# ---------------------------------------------------------------------------
# 全局样式
# ---------------------------------------------------------------------------

st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600&display=swap');
    html, body, [class*="css"] {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
        font-weight: 400;
        color: #1F2937;
        background-color: #F5F6FA;
    }
    .stButton > button {
        height: 44px;
        border-radius: 10px;
        background: linear-gradient(135deg, #5B5FFF 0%, #7A5CFF 100%);
        color: white;
        font-weight: 500;
        border: none;
        transition: all 0.2s ease;
        padding: 0 24px;
    }
    .stButton > button:hover {
        background: linear-gradient(135deg, #4C4CFF 0%, #6A4FFF 100%);
        transform: translateY(-1px);
        box-shadow: 0 10px 25px rgba(91,95,255,0.3);
    }
    .stButton > button:active {
        transform: translateY(0);
    }
    .stDownloadButton > button {
        height: 44px;
        border-radius: 10px;
        background: linear-gradient(135deg, #5B5FFF 0%, #7A5CFF 100%);
        color: white;
        font-weight: 500;
        border: none;
        transition: all 0.2s ease;
    }
    .stDownloadButton > button:hover {
        background: linear-gradient(135deg, #4C4CFF 0%, #6A4FFF 100%);
        transform: translateY(-1px);
        box-shadow: 0 10px 25px rgba(91,95,255,0.3);
    }
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea {
        border-radius: 10px;
        border: 1px solid #E5E7EB;
        padding: 12px;
        background: #FFFFFF;
        font-size: 14px;
    }
    .stTextInput > div > div > input:focus,
    .stTextArea > div > div > textarea:focus {
        border-color: #5B5FFF;
        box-shadow: 0 0 0 3px rgba(91,95,255,0.15);
    }
    [data-testid="stExpander"] {
        background: #FFFFFF;
        border-radius: 16px;
        box-shadow: 0px 10px 30px rgba(0, 0, 0, 0.08), 0px 2px 6px rgba(0, 0, 0, 0.04);
        border: none;
    }
    .block-container {
        padding: 2rem 3rem;
        max-width: 900px;
    }
    [data-testid="stSidebar"] {
        display: none;
    }
    .looklook-title {
        text-align: center;
        font-size: 42px;
        font-weight: 600;
        margin: 0 0 0.25rem 0;
        background: linear-gradient(135deg, #5B5FFF 0%, #7A5CFF 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        background-clip: text;
    }
    .looklook-subtitle {
        text-align: center;
        font-size: 16px;
        color: #6B7280;
        margin: 0 0 2rem 0;
    }
    .summary-card {
        background: #F3F4F6;
        border-radius: 16px;
        padding: 1.25rem 1.5rem;
        margin: 1.5rem 0;
        line-height: 1.7;
    }
    .summary-card h4 {
        margin: 0 0 0.75rem 0;
        color: #1F2937;
    }
    .divider-line {
        border: none;
        border-top: 1px solid #E5E7EB;
        margin: 1.5rem 0;
    }
    .center-btn-wrap {
        display: flex;
        justify-content: center;
        margin-top: 1.5rem;
    }
    .center-btn-wrap .stButton > button {
        width: 200px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

# ---------------------------------------------------------------------------
# Session state 初始化
# ---------------------------------------------------------------------------

if "page" not in st.session_state:
    st.session_state.page = "input"
if "is_running" not in st.session_state:
    st.session_state.is_running = False
if "excel_bytes" not in st.session_state:
    st.session_state.excel_bytes = None
if "excel_filename" not in st.session_state:
    st.session_state.excel_filename = None
if "result_summary" not in st.session_state:
    st.session_state.result_summary = None
if "selected_papers" not in st.session_state:
    st.session_state.selected_papers = None
if "other_count" not in st.session_state:
    st.session_state.other_count = 0
if "search_params" not in st.session_state:
    st.session_state.search_params = None
if "pipeline_steps" not in st.session_state:
    st.session_state.pipeline_steps = []


# ---------------------------------------------------------------------------
# 工具函数
# ---------------------------------------------------------------------------


def get_date_range(time_label: str) -> tuple[date, date]:
    """根据时间范围标签计算 start_date 和 end_date。"""
    today = date.today()
    days = TIME_RANGE_DAYS[time_label]
    start = today - timedelta(days=days)
    return start, today


def call_deepseek(
    api_key: str,
    messages: list[dict[str, str]],
    *,
    json_mode: bool = True,
    temperature: float = 0.3,
) -> str:
    """调用 DeepSeek Chat API，带重试与错误处理。"""
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    body: dict[str, Any] = {
        "model": DEEPSEEK_MODEL,
        "messages": messages,
        "temperature": temperature,
    }
    if json_mode:
        body["response_format"] = {"type": "json_object"}

    last_error: Exception | None = None
    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.post(DEEPSEEK_API_URL, headers=headers, json=body, timeout=120)
            if resp.status_code == 401:
                raise ValueError("DeepSeek API Key 无效或已过期，请检查后重试。")
            if resp.status_code == 429:
                raise ValueError("DeepSeek API 请求过于频繁，请稍后再试。")
            if resp.status_code >= 400:
                raise RuntimeError(
                    f"DeepSeek API 错误 (HTTP {resp.status_code}): {resp.text[:300]}"
                )
            data = resp.json()
            content = data["choices"][0]["message"]["content"]
            return content.strip()
        except ValueError:
            raise
        except Exception as e:
            last_error = e
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_INTERVAL)
    raise RuntimeError(f"DeepSeek API 调用失败（已重试 {MAX_RETRIES} 次）: {last_error}")


def parse_json_response(text: str) -> Any:
    """解析 DeepSeek 返回的 JSON，支持外层包裹对象。"""
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    parsed = json.loads(text)
    if isinstance(parsed, dict):
        for key in ("results", "papers", "items", "data", "relevant_papers", "batch"):
            if key in parsed and isinstance(parsed[key], list):
                return parsed[key]
        for v in parsed.values():
            if isinstance(v, list):
                return v
    return parsed


def expand_keywords(
    api_key: str, keywords: str, interest: str, journal_input: str
) -> tuple[list[str], str]:
    """关键词拓展、期刊翻译与 OpenAlex 广撒网搜索词生成。"""
    keywords = keywords.strip()
    journal_hint = journal_input.strip() if journal_input.strip() else "（用户未填写期刊）"
    keywords_note = (
        "（用户未填写关键词，请完全依据个人兴趣描述生成 simple_terms）"
        if not keywords
        else keywords
    )

    system_msg = (
        "你是学术文献检索专家，专门为 OpenAlex 设计「广撒网」式检索词。"
        "目标是尽可能多地召回可能相关的论文，因此搜索词要宽泛，包含同义词和近义词，"
        "不要使用双引号做短语精确匹配。"
        "必须返回合法 JSON，且只包含 simple_terms 和 journal_names 两个字段。"
        "无论用户是否提供关键词，都必须返回 simple_terms 字段。"
    )
    user_msg = f"""请根据以下信息返回 JSON 对象（仅 JSON，无其他文字）：

字段要求：
1. "simple_terms": 用于 OpenAlex search 参数的英文检索串。策略：广撒网、高召回。
   - 使用 3–5 个最核心的英文关键词，以空格分隔
   - 不要使用双引号包裹短语
   - 同义词、近义词用大写 OR 连接，例如：bias OR prejudice interpersonal OR social relationships
   - 若兴趣涉及多个维度，可用 OR 连接不同维度
   - 词组不要加引号；OR 必须大写
   - **若用户关键词为空**，则完全依据「个人兴趣描述」自动生成 simple_terms，
     范围要广泛，涵盖同义词、近义词，使用 OR 连接不同概念，确保 OpenAlex 能尽可能多地召回相关论文
2. "journal_names": 字符串数组。若用户填写了期刊（中英文均可），翻译为标准英文期刊全称；
   若未填写则返回空数组 []

用户关键词：{keywords_note}
个人兴趣描述：{interest}
用户期刊输入：{journal_hint}

返回格式示例：
{{"simple_terms": "bias OR prejudice interpersonal relationships", "journal_names": ["Nature"]}}"""

    content = call_deepseek(
        api_key,
        [{"role": "system", "content": system_msg}, {"role": "user", "content": user_msg}],
        json_mode=True,
    )
    data = json.loads(content)
    journal_names = data.get("journal_names", [])
    simple_terms = data.get("simple_terms", "")
    if not isinstance(journal_names, list):
        journal_names = []
    if not simple_terms or not str(simple_terms).strip():
        simple_terms = keywords if keywords else interest[:80]
    return [str(j) for j in journal_names], str(simple_terms).strip()


def fetch_papers_openalex(
    search_terms: str,
    start_date: str,
    end_date: str,
    journal_names: list[str],
    progress_callback=None,
    max_papers: int = MAX_FETCH_PAPERS,
) -> list[dict]:
    """使用 OpenAlex API 抓取论文。"""
    papers: list[dict] = []
    page = 1

    while True:
        params = {
            "search": search_terms,
            "filter": f"from_publication_date:{start_date},to_publication_date:{end_date}",
            "per_page": OPENALEX_PER_PAGE,
            "page": page,
        }

        if journal_names and page == 1:
            source_ids = []
            for jname in journal_names:
                jname = jname.strip()
                if len(jname) <= 1:
                    continue
                try:
                    source_url = "https://api.openalex.org/sources"
                    source_params = {"search": jname, "per_page": 3}
                    source_resp = None
                    for sa in range(3):
                        try:
                            source_resp = requests.get(
                                source_url, params=source_params, timeout=15
                            )
                            break
                        except requests.exceptions.ConnectionError:
                            if sa < 2:
                                time.sleep(2)
                            else:
                                source_resp = None
                    if source_resp is None or source_resp.status_code != 200:
                        continue
                    source_data = source_resp.json()
                    for src in source_data.get("results", []):
                        sid = src.get("id", "")
                        if sid:
                            short_id = sid.split("/")[-1] if "/" in sid else sid
                            source_ids.append(short_id)
                            break
                except Exception:
                    pass

            if source_ids:
                if len(source_ids) == 1:
                    source_filter = source_ids[0]
                else:
                    source_filter = "|".join(source_ids)
                existing_filter = params.get("filter", "")
                if existing_filter:
                    params["filter"] = (
                        existing_filter + f",primary_location.source.id:{source_filter}"
                    )
                else:
                    params["filter"] = f"primary_location.source.id:{source_filter}"

        resp = None
        data = None
        max_retries = 3
        retry_delay = 3
        for attempt in range(max_retries):
            try:
                resp = requests.get(OPENALEX_SEARCH_URL, params=params, timeout=30)
                resp.raise_for_status()
                data = resp.json()
                break
            except requests.exceptions.ConnectionError as ce:
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    continue
                raise RuntimeError(f"OpenAlex 连接失败（已重试{max_retries}次）: {ce}")
            except Exception as e:
                if resp is not None:
                    raise RuntimeError(
                        f"OpenAlex 请求失败 (HTTP {resp.status_code}): {resp.text[:300]}"
                    ) from e
                raise RuntimeError(f"OpenAlex 请求失败: {e}") from e

        for work in data.get("results", []):
            if len(papers) >= max_papers:
                break
            abstract = ""
            inverted_index = work.get("abstract_inverted_index")
            if inverted_index:
                try:
                    max_pos = max(max(positions) for positions in inverted_index.values())
                    words = [None] * (max_pos + 1)
                    for word, positions in inverted_index.items():
                        for pos in positions:
                            words[pos] = word
                    abstract = " ".join(filter(None, words))
                except Exception:
                    pass

            authors = ", ".join(
                [
                    a["author"]["display_name"]
                    for a in work.get("authorships", [])
                    if a.get("author")
                ]
            )
            try:
                journal = (
                    work.get("primary_location", {})
                    .get("source", {})
                    .get("display_name", "")
                )
            except AttributeError:
                journal = ""
            year = work.get("publication_year", "")
            citation_count = work.get("cited_by_count", 0)

            papers.append(
                {
                    "title": work.get("title", ""),
                    "authors": authors,
                    "abstract": abstract,
                    "journal": journal,
                    "citationCount": citation_count,
                    "year": int(year) if year else 0,
                    "paperId": work.get("id", ""),
                }
            )

        if progress_callback:
            progress_callback(len(papers))

        total_count = data.get("meta", {}).get("count", 0)
        if (
            len(papers) >= max_papers
            or len(papers) >= total_count
            or len(data.get("results", [])) == 0
        ):
            break
        page += 1
        time.sleep(OPENALEX_SLEEP)

    return papers


def filter_papers_batch(
    api_key: str, batch: list[dict], interest: str, batch_num: int
) -> list[dict]:
    """AI 智能筛选一批论文。"""
    lines = []
    for i, p in enumerate(batch):
        abstract = p.get("abstract") or "（无摘要）"
        lines.append(f"[{i}] 标题: {p['title']}\n摘要: {abstract[:800]}")
    papers_text = "\n\n".join(lines)

    system_msg = (
        "你是学术论文筛选助手。根据用户兴趣判断每篇论文是否高度相关。"
        "必须返回 JSON 对象，包含 results 数组。"
    )
    user_msg = f"""用户个人兴趣描述：
{interest}

以下论文（批次内序号从 0 开始）：
{papers_text}

请返回 JSON：
{{"results": [{{"index": 0, "relevant": true}}, {{"index": 1, "relevant": false}}, ...]}}

对每篇论文判断 relevant（true/false）。仅返回 JSON。"""

    content = call_deepseek(
        api_key,
        [{"role": "system", "content": system_msg}, {"role": "user", "content": user_msg}],
        json_mode=True,
    )
    parsed = json.loads(content)
    results = parsed.get("results", parsed) if isinstance(parsed, dict) else parsed
    if not isinstance(results, list):
        results = parse_json_response(content)
        if isinstance(results, dict):
            results = list(results.values())[0] if results else []

    relevant_indices = set()
    for item in results:
        if isinstance(item, dict) and item.get("relevant") is True:
            relevant_indices.add(item.get("index", -1))

    return [batch[i] for i in range(len(batch)) if i in relevant_indices]


def enrich_papers_batch(api_key: str, batch: list[dict], interest: str) -> list[dict]:
    """为一批论文生成中文信息与评分。"""
    lines = []
    for i, p in enumerate(batch):
        abstract = p.get("abstract") or "（无摘要）"
        lines.append(f"[{i}] 标题: {p['title']}\n摘要: {abstract[:600]}")

    system_msg = "你是学术论文分析助手。为每篇论文生成中文翻译、总结和评分。必须返回 JSON。"
    user_msg = f"""用户兴趣：{interest}

论文列表：
{chr(10).join(lines)}

返回 JSON：
{{"results": [
  {{
    "index": 0,
    "relevance_score": 0.85,
    "title_cn": "中文标题",
    "one_liner": "一句话总结（中文）",
    "recommendation_reason": "推荐理由（中文，约30字）"
  }}
]}}

relevance_score 为 0-1 浮点数。仅返回 JSON。"""

    content = call_deepseek(
        api_key,
        [{"role": "system", "content": system_msg}, {"role": "user", "content": user_msg}],
        json_mode=True,
    )
    parsed = json.loads(content)
    results = parsed.get("results", [])
    if not isinstance(results, list):
        results = parse_json_response(content)
        if not isinstance(results, list):
            results = []

    enrich_map: dict[int, dict] = {}
    for item in results:
        if isinstance(item, dict) and "index" in item:
            enrich_map[item["index"]] = item

    enriched = []
    for i, p in enumerate(batch):
        info = enrich_map.get(i, {})
        paper = dict(p)
        paper["relevance_score"] = float(info.get("relevance_score", 0.5))
        paper["title_cn"] = info.get("title_cn") or p["title"]
        paper["one_liner"] = info.get("one_liner", "")
        paper["recommendation_reason"] = info.get("recommendation_reason", "")
        enriched.append(paper)
    return enriched


def compute_scores_and_select(
    candidate_papers: list[dict],
) -> tuple[list[dict], list[dict]]:
    """计算推荐指数、排序并分为精选与其他。"""
    counts = [p.get("citationCount", 0) or 0 for p in candidate_papers]
    min_c, max_c = min(counts), max(counts)
    if max_c == min_c:
        norm_citations = [1.0] * len(counts)
    else:
        norm_citations = [(c - min_c) / (max_c - min_c) for c in counts]

    for p, norm_c in zip(candidate_papers, norm_citations):
        rel = float(p.get("relevance_score", 0.5))
        p["norm_citation"] = norm_c
        p["score"] = round((rel * 0.5 + norm_c * 0.5) * 10, 1)

    sorted_papers = sorted(candidate_papers, key=lambda x: x["score"], reverse=True)
    selected = sorted_papers[:MAX_SELECTED_PAPERS]
    other = sorted_papers[MAX_SELECTED_PAPERS:]
    return selected, other


def generate_review(api_key: str, selected_papers: list[dict]) -> str:
    """生成中文综述。"""
    lines = []
    for p in selected_papers:
        lines.append(f"- {p.get('title_cn', p['title'])}：{p.get('one_liner', '')}")
    summary_input = "\n".join(lines)

    system_msg = "你是学术综述撰写专家，用流畅中文撰写文献速递综述。"
    user_msg = f"""根据以下精选论文的中文标题与一句话总结，撰写中文文献速递综述。
直接输出综述正文，不要额外标题。综述必须包含以下三部分：

1. 以「本期研究围绕……」开头，概括本期研究方向
2. 以「覆盖……方向」描述主要主题分布
3. 以「值得关注的结论有：」引出 5 条结论，每条以数字序号开头（1. 2. 3. 4. 5.）

总字数 200-400 字，语言流畅自然。

论文信息：
{summary_input}"""

    return call_deepseek(
        api_key,
        [{"role": "system", "content": system_msg}, {"role": "user", "content": user_msg}],
        json_mode=False,
        temperature=0.5,
    )


def build_excel(review: str, selected: list[dict], other: list[dict]) -> bytes:
    """生成三 Sheet 的 Excel 文件。"""
    wb = Workbook()
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")

    ws1 = wb.active
    ws1.title = "文献速递"
    ws1.merge_cells("A1:J1")
    cell = ws1["A1"]
    cell.value = review
    cell.alignment = wrap
    ws1.row_dimensions[1].height = 120
    ws1["A3"] = f"以上综述基于精选的 {len(selected)} 篇论文，详情见 Sheet2。"

    ws2 = wb.create_sheet("精选论文数据")
    headers2 = [
        "推荐指数",
        "英文标题",
        "中文标题",
        "作者",
        "期刊",
        "发表年份",
        "引用次数",
        "英文摘要",
        "一句话总结（中文）",
        "推荐理由（中文）",
    ]
    col_widths2 = [10, 30, 30, 20, 20, 8, 8, 50, 40, 40]
    for col, (h, w) in enumerate(zip(headers2, col_widths2), 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = bold
        ws2.column_dimensions[get_column_letter(col)].width = w

    for row_idx, p in enumerate(selected, 2):
        ws2.cell(row=row_idx, column=1, value=p.get("score", 0))
        ws2.cell(row=row_idx, column=2, value=p.get("title", ""))
        ws2.cell(row=row_idx, column=3, value=p.get("title_cn", ""))
        ws2.cell(row=row_idx, column=4, value=p.get("authors", ""))
        ws2.cell(row=row_idx, column=5, value=p.get("journal", ""))
        ws2.cell(row=row_idx, column=6, value=p.get("year", ""))
        ws2.cell(row=row_idx, column=7, value=p.get("citationCount", 0))
        c8 = ws2.cell(row=row_idx, column=8, value=p.get("abstract", ""))
        c8.alignment = wrap
        c9 = ws2.cell(row=row_idx, column=9, value=p.get("one_liner", ""))
        c9.alignment = wrap
        c10 = ws2.cell(row=row_idx, column=10, value=p.get("recommendation_reason", ""))
        c10.alignment = wrap

    ws3 = wb.create_sheet("其他相关论文")
    for col, h in enumerate(["英文标题", "中文标题"], 1):
        c = ws3.cell(row=1, column=col, value=h)
        c.font = bold
        ws3.column_dimensions[get_column_letter(col)].width = 40

    if other:
        for row_idx, p in enumerate(other, 2):
            ws3.cell(row=row_idx, column=1, value=p.get("title", ""))
            ws3.cell(row=row_idx, column=2, value=p.get("title_cn", ""))
    else:
        ws3["A2"] = "无"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def run_pipeline(
    api_key: str,
    keywords: str,
    journal_input: str,
    time_label: str,
    interest: str,
    status_container,
) -> bool:
    """主处理流程，成功返回 True。"""
    steps: list[str] = []
    start_date, end_date = get_date_range(time_label)
    start_date_str = start_date.strftime("%Y-%m-%d")
    end_date_str = end_date.strftime("%Y-%m-%d")

    status_container.update(label="🔍 正在拆解兴趣，拓展搜索词…", state="running")
    journal_names, simple_terms = expand_keywords(
        api_key, keywords, interest, journal_input
    )
    steps.append("✅ 正在拆解兴趣，拓展搜索词…")

    journal_names_for_fetch = journal_names if journal_input.strip() else []

    def on_fetch_progress(count: int) -> None:
        status_container.update(
            label=f"📡 正在从 OpenAlex 抓取论文（已获取 {count} 篇）…",
            state="running",
        )

    status_container.update(label="📡 正在从 OpenAlex 抓取论文…", state="running")
    raw_papers = fetch_papers_openalex(
        simple_terms,
        start_date_str,
        end_date_str,
        journal_names_for_fetch,
        progress_callback=on_fetch_progress,
    )
    steps.append(f"✅ 正在从 OpenAlex 抓取论文…（共 {len(raw_papers)} 篇）")

    if not raw_papers:
        status_container.update(label="未找到符合日期/期刊条件的论文", state="error")
        st.error("未找到符合日期或期刊条件的论文，请调整关键词或时间范围后重试。")
        st.session_state.pipeline_steps = steps
        return False

    candidate_papers: list[dict] = []
    batch_size_filter = 10
    total_batches = (len(raw_papers) + batch_size_filter - 1) // batch_size_filter
    for b in range(total_batches):
        status_container.update(
            label=f"🤖 AI 正在筛选论文（第 {b + 1}/{total_batches} 批）…",
            state="running",
        )
        batch = raw_papers[b * batch_size_filter : (b + 1) * batch_size_filter]
        filtered = filter_papers_batch(api_key, batch, interest, b + 1)
        candidate_papers.extend(filtered)
        time.sleep(0.5)
    steps.append(f"✅ AI 正在筛选…（保留 {len(candidate_papers)} 篇候选）")

    if not candidate_papers:
        status_container.update(label="未找到高度相关论文", state="error")
        st.warning("未找到高度相关论文，请调整兴趣描述后重试。")
        st.session_state.pipeline_steps = steps
        return False

    enriched_all: list[dict] = []
    batch_size_enrich = 5
    total_enrich = (len(candidate_papers) + batch_size_enrich - 1) // batch_size_enrich
    for b in range(total_enrich):
        status_container.update(
            label=f"⭐ 正在生成总结与评分（第 {b + 1}/{total_enrich} 批）…",
            state="running",
        )
        batch = candidate_papers[b * batch_size_enrich : (b + 1) * batch_size_enrich]
        enriched = enrich_papers_batch(api_key, batch, interest)
        enriched_all.extend(enriched)
        time.sleep(0.5)
    steps.append("✅ 正在生成总结与评分…")

    selected_papers, other_papers = compute_scores_and_select(enriched_all)

    status_container.update(label="📝 正在生成综述…", state="running")
    review = generate_review(api_key, selected_papers)
    steps.append("✅ 正在生成综述…")

    status_container.update(label="📊 正在生成 Excel…", state="running")
    excel_bytes = build_excel(review, selected_papers, other_papers)
    filename = f"文献速递_{date.today().strftime('%Y%m%d')}.xlsx"
    steps.append("✅ 正在生成 Excel…")

    journal_display = journal_input.strip() if journal_input.strip() else None
    if not journal_display and journal_names:
        journal_display = "、".join(journal_names)

    st.session_state.excel_bytes = excel_bytes
    st.session_state.excel_filename = filename
    st.session_state.result_summary = review
    st.session_state.selected_papers = selected_papers
    st.session_state.other_count = len(other_papers)
    st.session_state.pipeline_steps = steps
    st.session_state.search_params = {
        "interest": interest,
        "keywords": keywords.strip() if keywords.strip() else "（由 AI 自动生成）",
        "simple_terms": simple_terms,
        "time_label": time_label,
        "journal": journal_display,
    }

    status_container.update(label="✅ 分析完成！", state="complete")
    return True


# ---------------------------------------------------------------------------
# UI 组件
# ---------------------------------------------------------------------------


def render_looklook_header() -> None:
    st.markdown('<h1 class="looklook-title">LOOKLOOK</h1>', unsafe_allow_html=True)
    st.markdown(f'<p class="looklook-subtitle">{SUBTITLE}</p>', unsafe_allow_html=True)


def reset_to_input() -> None:
    st.session_state.page = "input"
    st.session_state.is_running = False
    st.session_state.excel_bytes = None
    st.session_state.excel_filename = None
    st.session_state.result_summary = None
    st.session_state.selected_papers = None
    st.session_state.other_count = 0
    st.session_state.search_params = None
    st.session_state.pipeline_steps = []


def render_input_page() -> None:
    render_looklook_header()

    _, center, _ = st.columns([1, 3, 1])
    with center:
        interest = st.text_area(
            "个人兴趣描述",
            height=120,
            placeholder="用通俗中文描述你感兴趣的研究方向，AI 将自动扩展搜索词",
            help="必填",
        )

        with st.expander("⚙️ 高级选项", expanded=False):
            keywords = st.text_input(
                "关键词（可选）",
                placeholder="多个关键词用逗号分隔，留空则由 AI 自动生成",
                label_visibility="visible",
            )
            journal_input = st.text_input(
                "期刊名称（可选）",
                placeholder="多个期刊用逗号分隔，留空则不限制期刊",
                label_visibility="visible",
            )
            time_label = st.selectbox(
                "时间范围",
                TIME_RANGE_OPTIONS,
                index=TIME_RANGE_OPTIONS.index(DEFAULT_TIME_RANGE),
            )
            api_key = st.text_input(
                "DeepSeek API Key",
                type="password",
                placeholder="在此输入你的 DeepSeek API Key，不会存储",
                label_visibility="visible",
            )

        st.markdown('<div class="center-btn-wrap">', unsafe_allow_html=True)
        start_btn = st.button("🚀 开始分析", key="start_analysis")
        st.markdown("</div>", unsafe_allow_html=True)

    if start_btn:
        if st.session_state.is_running:
            st.warning("正在处理中，请勿重复点击。")
            return

        if not api_key or not api_key.strip():
            st.error("请填写 DeepSeek API Key。")
            return
        if not interest or not interest.strip():
            st.error("请填写个人兴趣描述。")
            return

        st.session_state.excel_bytes = None
        st.session_state.excel_filename = None
        st.session_state.result_summary = None
        st.session_state.selected_papers = None
        st.session_state.other_count = 0
        st.session_state.search_params = None
        st.session_state.pipeline_steps = []
        st.session_state.is_running = True

        try:
            with st.status("正在启动分析…", expanded=True) as status:
                success = run_pipeline(
                    api_key.strip(),
                    keywords.strip() if keywords else "",
                    journal_input.strip() if journal_input else "",
                    time_label,
                    interest.strip(),
                    status,
                )
            if success:
                st.session_state.page = "result"
                st.rerun()
        except ValueError as e:
            st.error(str(e))
        except Exception as e:
            st.error(f"处理失败：{e}")
        finally:
            st.session_state.is_running = False


def render_result_page() -> None:
    top_l, top_r = st.columns([5, 1])
    with top_r:
        if st.button("← 重新搜索", key="back_to_search"):
            reset_to_input()
            st.rerun()

    render_looklook_header()
    st.markdown('<hr class="divider-line">', unsafe_allow_html=True)

    params = st.session_state.search_params or {}
    interest_text = html.escape(str(params.get("interest", "")))
    simple_terms = html.escape(str(params.get("simple_terms", "")))
    time_label = html.escape(str(params.get("time_label", "")))
    journal = params.get("journal")

    journal_line = ""
    if journal:
        journal_line = f"<p><strong>期刊名称：</strong>{html.escape(str(journal))}</p>"

    st.markdown(
        f"""
        <div class="summary-card">
            <h4>📌 本次检索摘要</h4>
            <p><strong>个人兴趣描述：</strong>{interest_text}</p>
            <p><strong>扩展关键词：</strong>{simple_terms}</p>
            <p><strong>时间范围：</strong>{time_label}</p>
            {journal_line}
        </div>
        """,
        unsafe_allow_html=True,
    )

    steps = st.session_state.pipeline_steps or []
    if steps:
        steps_text = "\n".join(f"- {s}" for s in steps)
        st.info(f"**处理进度**\n\n{steps_text}")

    st.success("✅ 分析完成！")

    selected = st.session_state.selected_papers or []
    review = st.session_state.result_summary or ""
    count = len(selected)

    st.subheader(f"📊 本次精选文献 {count} 条")
    if st.session_state.other_count:
        st.caption(f"另有 {st.session_state.other_count} 篇相关论文见 Excel Sheet3")

    st.markdown("### 📄 文献速递")
    st.markdown(review.replace("\n", "\n\n") if review else "")

    sorted_papers = sorted(selected, key=lambda x: x.get("score", 0), reverse=True)
    with st.expander("📋 精选论文一览", expanded=False):
        for i, p in enumerate(sorted_papers, 1):
            title_en = html.escape(str(p.get("title", "")))
            title_cn = html.escape(str(p.get("title_cn", "")))
            one_liner = html.escape(str(p.get("one_liner", "")))
            score = p.get("score", 0)
            seq = f"{i:02d}"
            block = f"""
<div style="margin-bottom:1.5em;">
  <div style="font-weight:bold;">{seq} {title_en}
    <span style="float:right;">推荐指数：{score}</span>
  </div>
  <div>{title_cn}</div>
  <div style="color:#555;">{one_liner}</div>
</div>
"""
            st.markdown(block, unsafe_allow_html=True)

    col1, col2, col3 = st.columns([1, 1, 1])
    with col3:
        st.download_button(
            label="📥 下载文献速递",
            data=st.session_state.excel_bytes,
            file_name=st.session_state.excel_filename
            or f"文献速递_{date.today().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="download_excel",
        )


# ---------------------------------------------------------------------------
# 主入口：按页面状态渲染
# ---------------------------------------------------------------------------

if st.session_state.page == "input":
    render_input_page()
elif st.session_state.page == "result":
    if st.session_state.result_summary and st.session_state.excel_bytes:
        render_result_page()
    else:
        st.session_state.page = "input"
        st.rerun()
