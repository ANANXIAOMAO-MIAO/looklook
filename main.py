"""
LOOKLOOK — FastAPI 后端：文献检索与智能速递流水线
"""

from __future__ import annotations

import base64
import hashlib
import json
import os
import re
import secrets
import sqlite3
import threading
import time
import uuid
from contextlib import contextmanager
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo
from io import BytesIO
from pathlib import Path
from typing import Any, Callable, Iterator

import psycopg
from psycopg.rows import dict_row
from psycopg_pool import ConnectionPool

import jwt
import requests
import uvicorn
from fastapi import FastAPI, Header, HTTPException, Query, Request
from fastapi.responses import FileResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from pydantic import BaseModel, Field

# ---------------------------------------------------------------------------
# 常量
# ---------------------------------------------------------------------------

FREE_MODE = True
DEFAULT_DEEPSEEK_KEY = os.getenv("DEFAULT_DEEPSEEK_KEY", "")

DEEPSEEK_API_URL = "https://api.deepseek.com/chat/completions"
DEEPSEEK_MODEL = "deepseek-chat"
OPENALEX_SEARCH_URL = "https://api.openalex.org/works"
OPENALEX_PER_PAGE = 200
OPENALEX_SLEEP = 0.1
MAX_FETCH_PAPERS = 200
MAX_SELECTED_PAPERS = 10
RELEVANCE_THRESHOLD = 0.55
MAX_RETRIES = 3
RETRY_INTERVAL = 2
JSON_MAX_TOKENS = 4096
P3_ENRICH_MAX_TOKENS = 8192
BATCH_SIZE_ENRICH = 3
ENRICH_FAILED_TITLE_CN = "（中文标题生成失败，请见上方英文标题）"

TIME_RANGE_OPTIONS = ["一月内", "半年内", "一年内", "三年内", "五年内"]

TIME_RANGE_DAYS = {
    "一月内": 30,
    "半年内": 182,
    "一年内": 365,
    "三年内": 1095,
    "五年内": 1825,
}

# 异步任务进度存储 {task_id: {step, message, done, cancelled, result, error}}
task_status: dict[str, dict[str, Any]] = {}
task_lock = threading.Lock()


class TaskCancelledError(Exception):
    """用户取消搜索任务。"""

# 数据库：生产用 DATABASE_URL（Supabase PostgreSQL），本地默认 SQLite 文件
DB_PATH = Path(__file__).parent / "looklook.db"
DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
JWT_SECRET_KEY = "looklook-jwt-secret-change-in-production"
JWT_ALGORITHM = "HS256"
JWT_EXPIRE_DAYS = 7
EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
SCRYPT_N = 2**14
SCRYPT_R = 8
SCRYPT_P = 1
SCRYPT_DKLEN = 64
SCRYPT_SALT_LEN = 16

ANONYMOUS_DAILY_LIMIT = 3
LOGGED_IN_DAILY_LIMIT = 20
USAGE_TIMEZONE = ZoneInfo("Asia/Shanghai")

# ---------------------------------------------------------------------------
# 数据库与用户认证
# ---------------------------------------------------------------------------

_db_pool: ConnectionPool | None = None
PG_POOL_MIN_SIZE = 2
PG_POOL_MAX_SIZE = 10


def _use_postgres() -> bool:
    url = DATABASE_URL
    if not url or url.startswith("sqlite:"):
        return False
    return url.startswith(("postgresql://", "postgres://"))


def _sql(query: str) -> str:
    """SQLite 使用 ? 占位符；PostgreSQL 使用 %s。"""
    return query.replace("?", "%s") if _use_postgres() else query


def _is_unique_violation(exc: BaseException) -> bool:
    if isinstance(exc, sqlite3.IntegrityError):
        return True
    return isinstance(exc, psycopg.errors.UniqueViolation)


def init_db_pool() -> None:
    """启动时创建 PostgreSQL 连接池（复用连接，降低远程库延迟）。"""
    global _db_pool
    if not _use_postgres() or _db_pool is not None:
        return
    _db_pool = ConnectionPool(
        conninfo=DATABASE_URL,
        min_size=PG_POOL_MIN_SIZE,
        max_size=PG_POOL_MAX_SIZE,
        kwargs={"row_factory": dict_row},
        timeout=30,
    )


def close_db_pool() -> None:
    global _db_pool
    if _db_pool is not None:
        _db_pool.close()
        _db_pool = None


@contextmanager
def get_db_connection() -> Iterator[Any]:
    if _use_postgres():
        if _db_pool is None:
            init_db_pool()
        assert _db_pool is not None
        with _db_pool.connection() as conn:
            yield conn
    else:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        try:
            yield conn
        finally:
            conn.close()


def _db_execute(conn: Any, query: str, params: tuple[Any, ...] = ()) -> Any:
    return conn.execute(_sql(query), params)


def init_db() -> None:
    if _use_postgres():
        statements = [
            """
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                email TEXT UNIQUE NOT NULL,
                username TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS search_history (
                id SERIAL PRIMARY KEY,
                user_id INTEGER NOT NULL REFERENCES users(id),
                interest_desc TEXT,
                keywords TEXT,
                simple_terms TEXT,
                time_range TEXT,
                journals TEXT,
                summary TEXT,
                created_at TIMESTAMPTZ DEFAULT CURRENT_TIMESTAMP
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS anonymous_usage (
                identifier TEXT NOT NULL,
                search_date DATE NOT NULL,
                count INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (identifier, search_date)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS user_usage (
                user_id INTEGER NOT NULL,
                search_date DATE NOT NULL,
                count INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (user_id, search_date)
            )
            """,
        ]
    else:
        statements = [
            """
            CREATE TABLE IF NOT EXISTS users (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                email TEXT UNIQUE NOT NULL,
                username TEXT NOT NULL,
                password_hash TEXT NOT NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS search_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                user_id INTEGER NOT NULL,
                interest_desc TEXT,
                keywords TEXT,
                simple_terms TEXT,
                time_range TEXT,
                journals TEXT,
                summary TEXT,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY(user_id) REFERENCES users(id)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS anonymous_usage (
                identifier TEXT NOT NULL,
                search_date DATE NOT NULL,
                count INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (identifier, search_date)
            )
            """,
            """
            CREATE TABLE IF NOT EXISTS user_usage (
                user_id INTEGER NOT NULL,
                search_date DATE NOT NULL,
                count INTEGER NOT NULL DEFAULT 0,
                PRIMARY KEY (user_id, search_date)
            )
            """,
        ]

    with get_db_connection() as conn:
        for stmt in statements:
            _db_execute(conn, stmt)
        conn.commit()

    backend = "PostgreSQL (DATABASE_URL)" if _use_postgres() else f"SQLite ({DB_PATH})"
    print(f"[LOOKLOOK] 数据库已初始化: {backend}")


def _today_cn() -> date:
    return datetime.now(USAGE_TIMEZONE).date()


def _date_str(d: date | None = None) -> str:
    return (d or _today_cn()).isoformat()


def _get_anonymous_used(identifier: str, search_date: str | None = None) -> int:
    with get_db_connection() as conn:
        row = _db_execute(
            conn,
            "SELECT count FROM anonymous_usage WHERE identifier = ? AND search_date = ?",
            (identifier, search_date or _date_str()),
        ).fetchone()
    return int(row["count"]) if row else 0


def _get_user_used(user_id: int, search_date: str | None = None) -> int:
    with get_db_connection() as conn:
        row = _db_execute(
            conn,
            "SELECT count FROM user_usage WHERE user_id = ? AND search_date = ?",
            (user_id, search_date or _date_str()),
        ).fetchone()
    return int(row["count"]) if row else 0


def consume_anonymous_usage(identifier: str) -> bool:
    """扣减匿名用户今日配额，有余量返回 True。"""
    identifier = (identifier or "").strip()
    if not identifier:
        return False
    today = _date_str()
    with get_db_connection() as conn:
        row = _db_execute(
            conn,
            "SELECT count FROM anonymous_usage WHERE identifier = ? AND search_date = ?",
            (identifier, today),
        ).fetchone()
        current = int(row["count"]) if row else 0
        if current >= ANONYMOUS_DAILY_LIMIT:
            return False
        if row:
            _db_execute(
                conn,
                "UPDATE anonymous_usage SET count = count + 1 WHERE identifier = ? AND search_date = ?",
                (identifier, today),
            )
        else:
            _db_execute(
                conn,
                "INSERT INTO anonymous_usage (identifier, search_date, count) VALUES (?, ?, 1)",
                (identifier, today),
            )
        conn.commit()
    return True


def consume_user_usage(user_id: int) -> bool:
    """扣减登录用户今日配额，有余量返回 True。"""
    today = _date_str()
    with get_db_connection() as conn:
        row = _db_execute(
            conn,
            "SELECT count FROM user_usage WHERE user_id = ? AND search_date = ?",
            (user_id, today),
        ).fetchone()
        current = int(row["count"]) if row else 0
        if current >= LOGGED_IN_DAILY_LIMIT:
            return False
        if row:
            _db_execute(
                conn,
                "UPDATE user_usage SET count = count + 1 WHERE user_id = ? AND search_date = ?",
                (user_id, today),
            )
        else:
            _db_execute(
                conn,
                "INSERT INTO user_usage (user_id, search_date, count) VALUES (?, ?, 1)",
                (user_id, today),
            )
        conn.commit()
    return True


def build_usage_snapshot(
    *,
    user_id: int | None,
    identifier: str | None,
) -> dict[str, Any]:
    if user_id is not None:
        used = _get_user_used(user_id)
        limit = LOGGED_IN_DAILY_LIMIT
        logged_in = True
    else:
        ident = (identifier or "").strip()
        used = _get_anonymous_used(ident) if ident else 0
        limit = ANONYMOUS_DAILY_LIMIT
        logged_in = False
    return {
        "limit": limit,
        "used": used,
        "remaining": max(0, limit - used),
        "logged_in": logged_in,
    }


def _resolve_anonymous_identifier(session_id: str | None, client_host: str | None) -> str:
    sid = (session_id or "").strip()
    if sid:
        return sid
    host = (client_host or "").strip()
    return host or "unknown"


def hash_password(password: str) -> str:
    salt = secrets.token_bytes(SCRYPT_SALT_LEN)
    pwd_hash = hashlib.scrypt(
        password.encode("utf-8"),
        salt=salt,
        n=SCRYPT_N,
        r=SCRYPT_R,
        p=SCRYPT_P,
        dklen=SCRYPT_DKLEN,
    )
    return f"{salt.hex()}:{pwd_hash.hex()}"


def verify_password(password: str, stored: str) -> bool:
    try:
        salt_hex, hash_hex = stored.split(":", 1)
        salt = bytes.fromhex(salt_hex)
        expected = bytes.fromhex(hash_hex)
    except (ValueError, TypeError):
        return False
    pwd_hash = hashlib.scrypt(
        password.encode("utf-8"),
        salt=salt,
        n=SCRYPT_N,
        r=SCRYPT_R,
        p=SCRYPT_P,
        dklen=SCRYPT_DKLEN,
    )
    return secrets.compare_digest(pwd_hash, expected)


def create_access_token(user_id: int, email: str) -> str:
    now = datetime.now(timezone.utc)
    payload = {
        "user_id": user_id,
        "email": email,
        "exp": now + timedelta(days=JWT_EXPIRE_DAYS),
        "iat": now,
    }
    return jwt.encode(payload, JWT_SECRET_KEY, algorithm=JWT_ALGORITHM)


def _normalize_email(email: str) -> str:
    return email.strip().lower()


def _is_valid_email(email: str) -> bool:
    return bool(EMAIL_PATTERN.match(email))


def _get_user_by_email(email: str) -> Any | None:
    with get_db_connection() as conn:
        row = _db_execute(
            conn,
            "SELECT id, email, username, password_hash FROM users WHERE email = ?",
            (email,),
        ).fetchone()
    return row


def _get_user_by_id(user_id: int) -> Any | None:
    with get_db_connection() as conn:
        return _db_execute(
            conn,
            "SELECT id, email, username FROM users WHERE id = ?",
            (user_id,),
        ).fetchone()


def _extract_bearer_token(authorization: str | None) -> str | None:
    if not authorization:
        return None
    parts = authorization.split(None, 1)
    if len(parts) != 2 or parts[0].lower() != "bearer":
        return None
    return parts[1].strip() or None


def _decode_user_from_token(token: str) -> dict[str, Any] | None:
    try:
        payload = jwt.decode(token, JWT_SECRET_KEY, algorithms=[JWT_ALGORITHM])
    except jwt.PyJWTError:
        return None
    user_id = payload.get("user_id")
    if user_id is None:
        return None
    return {"user_id": int(user_id), "email": str(payload.get("email", ""))}


def _optional_user_id_from_header(authorization: str | None) -> int | None:
    token = _extract_bearer_token(authorization)
    if not token:
        return None
    decoded = _decode_user_from_token(token)
    return decoded["user_id"] if decoded else None


def _require_user_id_from_header(authorization: str | None) -> int:
    token = _extract_bearer_token(authorization)
    if not token:
        raise HTTPException(status_code=401, detail="未登录或令牌无效")
    decoded = _decode_user_from_token(token)
    if not decoded:
        raise HTTPException(status_code=401, detail="登录已过期，请重新登录")
    user = _get_user_by_id(decoded["user_id"])
    if user is None:
        raise HTTPException(status_code=401, detail="用户不存在")
    return int(user["id"])


def save_search_history(user_id: int, req: SearchRequest, result: dict[str, Any]) -> None:
    journals = result.get("journals_used") or []
    journals_text = json.dumps(journals, ensure_ascii=False) if journals else ""
    with get_db_connection() as conn:
        _db_execute(
            conn,
            """
            INSERT INTO search_history (
                user_id, interest_desc, keywords, simple_terms,
                time_range, journals, summary
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                user_id,
                result.get("interest") or req.interest.strip(),
                req.keywords.strip() if req.keywords else "",
                result.get("simple_terms") or "",
                result.get("time_range") or req.time_range,
                journals_text,
                result.get("summary") or "",
            ),
        )
        conn.commit()


# ---------------------------------------------------------------------------
# 流水线核心函数（自 app.py 提取，无 Streamlit 依赖）
# ---------------------------------------------------------------------------


def get_date_range(time_label: str) -> tuple[date, date]:
    """根据时间范围标签计算 start_date 和 end_date。"""
    if time_label not in TIME_RANGE_DAYS:
        raise ValueError(f"无效的时间范围: {time_label}")
    today = date.today()
    days = TIME_RANGE_DAYS[time_label]
    start = today - timedelta(days=days)
    return start, today


def call_deepseek(
    api_key: str,
    messages: list[dict[str, str]],
    *,
    json_mode: bool = True,
    temperature: float = 0.3,
    max_tokens: int | None = None,
) -> str:
    """调用 DeepSeek Chat API，带重试与错误处理。"""
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    body: dict[str, Any] = {
        "model": DEEPSEEK_MODEL,
        "messages": messages,
        "temperature": temperature,
    }
    if max_tokens is not None:
        body["max_tokens"] = max_tokens
    if json_mode:
        body["response_format"] = {"type": "json_object"}

    last_error: Exception | None = None
    for attempt in range(MAX_RETRIES):
        try:
            resp = requests.post(DEEPSEEK_API_URL, headers=headers, json=body, timeout=120)
            if resp.status_code == 401:
                raise ValueError("DeepSeek API Key 无效或已过期，请检查后重试。")
            if resp.status_code == 429:
                raise ValueError("DeepSeek API 请求过于频繁，请稍后再试。")
            if resp.status_code >= 400:
                raise RuntimeError(
                    f"DeepSeek API 错误 (HTTP {resp.status_code}): {resp.text[:300]}"
                )
            data = resp.json()
            content = data["choices"][0]["message"]["content"]
            return content.strip()
        except ValueError:
            raise
        except Exception as e:
            last_error = e
            if attempt < MAX_RETRIES - 1:
                time.sleep(RETRY_INTERVAL)
    raise RuntimeError(f"DeepSeek API 调用失败（已重试 {MAX_RETRIES} 次）: {last_error}")


def parse_json_response(
    content: str,
    *,
    stage: str = "",
    context: str = "",
) -> Any:
    """
    安全解析 AI 返回的 JSON。
    只负责：清理 markdown 标记 → json.loads → 返回原始解析结果。
    不做任何类型判断或字段提取，把 dict/list/str 原样交给调用方处理。
    """
    if not content:
        return {}

    text = content.strip()

    # 去掉 AI 可能包裹的 ```json ... ``` 标记
    if text.startswith("```"):
        lines = text.split("\n")
        if lines[0].strip().startswith("```"):
            lines = lines[1:]
        if lines and lines[-1].strip() == "```":
            lines = lines[:-1]
        text = "\n".join(lines).strip()

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        stage_part = f" stage={stage}" if stage else ""
        ctx_part = f" {context}" if context else ""
        print(
            f"[JSON解析失败]{stage_part}{ctx_part} len={len(content)} "
            f"原始返回前200字符: {content[:200]}"
        )
        return {}


def safe_get_json_field(
    data: Any,
    field: str,
    default: Any = None,
    expected_type: type | None = None,
) -> Any:
    """
    安全地从解析后的 JSON 数据中提取字段。
    data: parse_json_response 的返回值
    field: 字段名
    default: 默认值
    expected_type: 期望的类型（如 list, str），如果实际类型不匹配则返回 default
    """
    if not isinstance(data, dict):
        return default
    value = data.get(field, default)
    if expected_type is not None and not isinstance(value, expected_type):
        return default
    return value


def _extract_results_list(parsed: Any) -> list:
    """从 P2/P3 等阶段的 JSON 中提取 results 列表。"""
    results = safe_get_json_field(parsed, "results", default=None, expected_type=list)
    if isinstance(results, list):
        return results
    if isinstance(parsed, list):
        return parsed
    return []


def _simple_terms_to_openalex_query(simple_terms_raw: Any, fallback: str) -> tuple[str, str]:
    """将 P1 返回的 simple_terms（数组或字符串）转为 OpenAlex 查询串与展示文本。"""
    terms: list[str] = []
    if isinstance(simple_terms_raw, list):
        terms = [str(t).strip() for t in simple_terms_raw if str(t).strip()]
    elif isinstance(simple_terms_raw, str) and simple_terms_raw.strip():
        raw = simple_terms_raw.strip()
        if " OR " in raw.upper():
            parts = re.split(r"\s+OR\s+", raw, flags=re.IGNORECASE)
            terms = [p.strip() for p in parts if p.strip()]
        else:
            terms = [t for t in raw.split() if t.strip()]

    if not terms:
        fb = fallback.strip()
        return fb, fb

    query = " OR ".join(terms)
    display = ", ".join(terms)
    return query, display


def expand_keywords(
    api_key: str, keywords: str, interest: str, journal_input: str
) -> dict[str, Any]:
    """P1：用户意图解析、检索策略与筛选标尺生成。"""
    keywords = keywords.strip()
    journal_hint = journal_input.strip() if journal_input.strip() else "（用户未填写期刊）"
    if keywords:
        keywords_note = f"用户自行提供的关键词：{keywords}"
    else:
        keywords_note = "用户未填写关键词。"

    system_msg = (
        "你是用户意图解析与学术检索策略专家。"
        "你的任务是根据用户的自然语言兴趣描述，完成三件事：\n"
        "1. 理解用户的产出意图（如内容创作、报告支撑、领域探索）；\n"
        "2. 提取核心概念，生成用于 OpenAlex 的英文搜索词（必须广泛且包含同义词近义词，不使用短语精确匹配）；\n"
        "3. 制定论文筛选的评分指南，告诉后续筛选阶段「什么样的论文在该意图下算高分」。\n\n"
        "你必须返回一个合法 JSON 对象，且只包含 JSON，无其他文字。JSON 包含以下字段：\n"
        "  - intent_type: 字符串，取值为 \"content_creation\"（内容创作，如科普、小红书帖子）、"
        "\"report_support\"（报告/课题支撑，需引用文献论证观点）、\"exploration\"（领域探索，无明确产出要求）。\n"
        "  - core_concepts: 字符串，用中文简述用户的核心关注点，包括隐含的限定条件（物种、人群、方法偏好等），"
        "供后续筛选和综述生成使用。\n"
        "  - simple_terms: 数组，英文检索词。每个词应是一个单词或简短的 2-3 词短语（但不要加引号做精确匹配）。"
        "必须广泛撒网，包含同义词、近义词、上下位词。如果用户提到了具体的物种、人群、地域，"
        "必须将这些作为检索词加入（如 feline, cat, adolescent, Chinese）。\n"
        "  - journal_names: 数组，若用户指定期刊或你需要限定高影响期刊则列出 ISSN 或简称，否则为空数组。\n"
        "  - exclusion_terms: 数组，需要排除的英文术语。用于过滤明显无关方向的论文。"
        "必须基于用户意图和隐含限定生成，例如用户研究猫，则排除 human, rat, mouse 等；"
        "用户研究青少年，则排除 children, infant 等。每个词要确保是全文检索会匹配到的无关词汇。\n"
        "  - scoring_guide: 字符串，用中文说明在该用户意图下，什么样的论文应被 P2 评为高分。"
        "包含具体标准，如「是否有清晰的结论」「是否容易用通俗语言转述」「是否是综述或元分析」「引用量高低」等。"
        "这个指南将被用于筛选论文，所以必须可操作。\n\n"
        "重要原则：\n"
        "- 必须识别并保留用户的隐含限定词（尤其物种、人群、地域），绝不能丢失。\n"
        "- 如果用户描述看起来与医学/人类相关，但提到了动物，必须将动物词加入检索词并排除人类相关术语。\n"
        "- 搜索词数量建议 5-15 个，覆盖不同角度。\n"
        "- 排除词只添加那些肯定无关且会大量误召回的词，不要过度排除。\n"
        "- 思维链：先在内部简要分析用户意图、核心概念、限定条件，再生成 JSON。但最终只输出 JSON。"
    )
    user_msg = f"""用户兴趣描述：
{interest}

{keywords_note}
用户期刊输入：{journal_hint}

请返回 JSON 对象（仅 JSON，无其他文字）：
{{
  "intent_type": "...",
  "core_concepts": "...",
  "simple_terms": [...],
  "journal_names": [...],
  "exclusion_terms": [...],
  "scoring_guide": "..."
}}

重要：如果用户提供了关键词，你必须将这些关键词翻译/转化为适当的英文检索词，并确保它们出现在 simple_terms 数组中。"""

    content = call_deepseek(
        api_key,
        [{"role": "system", "content": system_msg}, {"role": "user", "content": user_msg}],
        json_mode=True,
        max_tokens=JSON_MAX_TOKENS,
    )
    data = parse_json_response(content, stage="P0")
    if not isinstance(data, dict):
        data = {}

    journal_names = safe_get_json_field(data, "journal_names", default=[], expected_type=list)

    exclusion_terms = safe_get_json_field(data, "exclusion_terms", default=[], expected_type=list)
    exclusion_terms = [
        str(t).strip().lower() for t in exclusion_terms if str(t).strip()
    ]

    simple_terms_raw = data.get("simple_terms", "")
    if not isinstance(simple_terms_raw, (list, str)):
        simple_terms_raw = ""

    fallback = keywords if keywords else interest[:80]
    simple_terms_query, simple_terms_display = _simple_terms_to_openalex_query(
        simple_terms_raw, fallback
    )

    core_concepts = str(safe_get_json_field(data, "core_concepts", default="") or "").strip()
    scoring_guide = str(safe_get_json_field(data, "scoring_guide", default="") or "").strip()
    intent_type = str(
        safe_get_json_field(data, "intent_type", default="exploration") or "exploration"
    ).strip()

    return {
        "journal_names": [str(j) for j in journal_names],
        "simple_terms": simple_terms_query,
        "simple_terms_display": simple_terms_display,
        "core_concepts": core_concepts,
        "exclusion_terms": exclusion_terms,
        "scoring_guide": scoring_guide,
        "intent_type": intent_type,
    }


def filter_papers_by_exclusion(
    papers: list[dict], exclusion_terms: list[str]
) -> list[dict]:
    """根据 P1 排除词过滤 OpenAlex 论文（标题+摘要命中任一词则丢弃）。"""
    if not exclusion_terms:
        return papers
    kept: list[dict] = []
    for paper in papers:
        text = f"{paper.get('title', '')} {paper.get('abstract', '')}".lower()
        if any(term in text for term in exclusion_terms):
            continue
        kept.append(paper)
    return kept


def fetch_papers_openalex(
    search_terms: str,
    start_date: str,
    end_date: str,
    journal_names: list[str],
    progress_callback: Callable[[int], None] | None = None,
    max_papers: int = MAX_FETCH_PAPERS,
    cancel_check: Callable[[], bool] | None = None,
) -> list[dict]:
    """使用 OpenAlex API 抓取论文。"""
    papers: list[dict] = []
    page = 1

    while True:
        if cancel_check and cancel_check():
            raise TaskCancelledError("搜索已取消")

        params = {
            "search": search_terms,
            "filter": f"from_publication_date:{start_date},to_publication_date:{end_date}",
            "per_page": OPENALEX_PER_PAGE,
            "page": page,
        }

        if journal_names and page == 1:
            source_ids = []
            for jname in journal_names:
                jname = jname.strip()
                if len(jname) <= 1:
                    continue
                try:
                    source_url = "https://api.openalex.org/sources"
                    source_params = {"search": jname, "per_page": 3}
                    source_resp = None
                    for sa in range(3):
                        try:
                            source_resp = requests.get(
                                source_url, params=source_params, timeout=15
                            )
                            break
                        except requests.exceptions.ConnectionError:
                            if sa < 2:
                                time.sleep(2)
                            else:
                                source_resp = None
                    if source_resp is None or source_resp.status_code != 200:
                        continue
                    source_data = source_resp.json()
                    for src in source_data.get("results", []):
                        sid = src.get("id", "")
                        if sid:
                            short_id = sid.split("/")[-1] if "/" in sid else sid
                            source_ids.append(short_id)
                            break
                except Exception:
                    pass

            if source_ids:
                if len(source_ids) == 1:
                    source_filter = source_ids[0]
                else:
                    source_filter = "|".join(source_ids)
                existing_filter = params.get("filter", "")
                if existing_filter:
                    params["filter"] = (
                        existing_filter + f",primary_location.source.id:{source_filter}"
                    )
                else:
                    params["filter"] = f"primary_location.source.id:{source_filter}"

        resp = None
        data = None
        max_retries = 3
        retry_delay = 3
        for attempt in range(max_retries):
            try:
                resp = requests.get(OPENALEX_SEARCH_URL, params=params, timeout=30)
                resp.raise_for_status()
                data = resp.json()
                break
            except requests.exceptions.ConnectionError as ce:
                if attempt < max_retries - 1:
                    time.sleep(retry_delay)
                    continue
                raise RuntimeError(f"OpenAlex 连接失败（已重试{max_retries}次）: {ce}")
            except Exception as e:
                if resp is not None:
                    raise RuntimeError(
                        f"OpenAlex 请求失败 (HTTP {resp.status_code}): {resp.text[:300]}"
                    ) from e
                raise RuntimeError(f"OpenAlex 请求失败: {e}") from e

        for work in data.get("results", []):
            if len(papers) >= max_papers:
                break
            abstract = ""
            inverted_index = work.get("abstract_inverted_index")
            if inverted_index:
                try:
                    max_pos = max(max(positions) for positions in inverted_index.values())
                    words = [None] * (max_pos + 1)
                    for word, positions in inverted_index.items():
                        for pos in positions:
                            words[pos] = word
                    abstract = " ".join(filter(None, words))
                except Exception:
                    pass

            authors = ", ".join(
                [
                    a["author"]["display_name"]
                    for a in work.get("authorships", [])
                    if a.get("author")
                ]
            )
            try:
                journal = (
                    work.get("primary_location", {})
                    .get("source", {})
                    .get("display_name", "")
                )
            except AttributeError:
                journal = ""
            year = work.get("publication_year", "")
            citation_count = work.get("cited_by_count", 0)

            papers.append(
                {
                    "title": work.get("title", ""),
                    "authors": authors,
                    "abstract": abstract,
                    "journal": journal,
                    "citationCount": citation_count,
                    "year": int(year) if year else 0,
                    "paperId": work.get("id", ""),
                }
            )

        if progress_callback:
            progress_callback(len(papers))

        total_count = data.get("meta", {}).get("count", 0)
        if (
            len(papers) >= max_papers
            or len(papers) >= total_count
            or len(data.get("results", [])) == 0
        ):
            break
        page += 1
        time.sleep(OPENALEX_SLEEP)

    print(f" 实际抓取到的论文数: {len(papers)}")
    return papers


def _tier_from_score(score: float) -> str:
    if score >= 0.75:
        return "high"
    if score >= RELEVANCE_THRESHOLD:
        return "medium"
    return "low"


def _intent_type_cn(intent_type: str) -> str:
    return {
        "content_creation": "内容创作",
        "report_support": "报告/课题支撑",
        "exploration": "领域探索",
    }.get(intent_type, "领域探索")


def filter_papers_batch(
    api_key: str,
    batch: list[dict],
    interest: str,
    batch_num: int,
    *,
    core_concepts: str = "",
    scoring_guide: str = "",
    intent_type: str = "exploration",
) -> list[dict]:
    """P2：评估相关度分数与分级，保留 score >= RELEVANCE_THRESHOLD 的论文。"""
    lines = []
    for i, p in enumerate(batch):
        abstract = p.get("abstract") or "（无摘要）"
        lines.append(f"[{i}] 标题: {p['title']}\n摘要: {abstract[:800]}")
    papers_text = "\n\n".join(lines)

    context_block = ""
    if core_concepts or scoring_guide:
        context_block = (
            f"\n\n用户核心关注点：{core_concepts}\n\n评分指南：{scoring_guide}"
        )

    system_msg = (
        "你是学术论文筛选助手。根据用户兴趣与评分指南，阅读每篇论文的标题与摘要，"
        "给出 0～1 的相关度分数与分级。\n"
        "规则：\n"
        "1. 只依据提供的标题与摘要判断，禁止编造未出现的信息。\n"
        "2. 严格按 scoring_guide 理解「在该用户意图下何谓高分」。\n"
        "3. relevance_score：1.0=高度契合；0.55=勉强相关下限；<0.55=应淘汰。\n"
        "4. tier：high(≥0.75)、medium(0.55～0.74)、low(<0.55)。\n"
        "5. relevant：relevance_score ≥ 0.55 时为 true，否则 false。\n"
        "必须返回 JSON：{\"results\": [{\"index\": 0, \"relevance_score\": 0.82, "
        "\"tier\": \"high\", \"relevant\": true}, ...]}。仅返回 JSON，不要其他文字。"
    )
    user_msg = f"""用户个人兴趣描述：
{interest}
用户产出意图：{intent_type}（{_intent_type_cn(intent_type)}）{context_block}

请对下列论文逐篇评分。边界篇宁可给低分，不要把明显无关或仅关键词沾边的论文标为 medium/high。

以下论文（批次内序号从 0 开始）：
{papers_text}

请返回 JSON：
{{"results": [{{"index": 0, "relevance_score": 0.82, "tier": "high", "relevant": true}}, ...]}}

仅返回 JSON。"""

    content = call_deepseek(
        api_key,
        [{"role": "system", "content": system_msg}, {"role": "user", "content": user_msg}],
        json_mode=True,
        max_tokens=JSON_MAX_TOKENS,
    )
    parsed = parse_json_response(content, stage="P2", context=f"batch={batch_num}")
    results = _extract_results_list(parsed)

    kept: list[dict] = []
    for item in results:
        if not isinstance(item, dict):
            continue
        try:
            idx = int(item.get("index", -1))
        except (TypeError, ValueError):
            continue
        if idx < 0 or idx >= len(batch):
            continue
        try:
            score = float(item.get("relevance_score", 0))
        except (TypeError, ValueError):
            if item.get("relevant") is True:
                score = RELEVANCE_THRESHOLD
            else:
                score = 0.0
        score = max(0.0, min(1.0, score))
        if score < RELEVANCE_THRESHOLD:
            continue
        tier = str(item.get("tier") or _tier_from_score(score)).strip().lower()
        if tier not in ("high", "medium", "low"):
            tier = _tier_from_score(score)
        paper = dict(batch[idx])
        paper["relevance_score"] = round(score, 4)
        paper["relevance_tier"] = tier
        kept.append(paper)

    return kept


def _paper_enrich_ok(paper: dict) -> bool:
    """判断 P3 中文材料是否生成完整。"""
    title_cn = str(paper.get("title_cn") or "").strip()
    if not title_cn or title_cn == ENRICH_FAILED_TITLE_CN:
        return False
    if title_cn == str(paper.get("title") or "").strip():
        return False
    summary = str(paper.get("one_liner") or paper.get("summary_cn") or "").strip()
    return bool(summary)


def _apply_enrich_info(paper: dict, info: dict) -> dict:
    """将单篇 AI  enrich 结果合并进论文字典。"""
    out = dict(paper)
    title_cn = str(info.get("title_cn") or "").strip()
    if title_cn and title_cn != str(paper.get("title") or "").strip():
        out["title_cn"] = title_cn
    else:
        out["title_cn"] = ENRICH_FAILED_TITLE_CN

    summary_cn = str(info.get("summary_cn") or info.get("one_liner") or "").strip()
    out["summary_cn"] = summary_cn
    out["one_liner"] = summary_cn

    recommendation_text = str(info.get("recommendation_text", "") or "").strip()
    if not recommendation_text:
        rec_raw = info.get("recommendation")
        if isinstance(rec_raw, dict):
            if rec_raw.get("short"):
                recommendation_text = str(rec_raw.get("short", "")).strip()
            else:
                parts = [
                    str(rec_raw.get("usage", "") or "").strip(),
                    str(rec_raw.get("evidence_level", "") or "").strip(),
                    str(rec_raw.get("highlight", "") or "").strip(),
                ]
                recommendation_text = "，".join(part for part in parts if part)
        if not recommendation_text:
            recommendation_text = str(info.get("recommendation_reason", "") or "").strip()
    out["recommendation_text"] = recommendation_text
    out["highlights"] = str(info.get("highlights", "") or "").strip()
    out["limitations"] = str(info.get("limitations", "") or "").strip()
    return out


def _enrich_papers_batch_once(
    api_key: str,
    batch: list[dict],
    interest: str,
    *,
    core_concepts: str = "",
    intent_type: str = "exploration",
    log_context: str = "",
) -> list[dict]:
    """P3 单次 API 调用：生成中文标题、通俗摘要与深度点评。"""
    papers_for_ai = []
    for i, p in enumerate(batch):
        papers_for_ai.append(
            {
                "id": i,
                "title": p.get("title", ""),
                "abstract": (p.get("abstract") or "（无摘要）")[:600],
            }
        )
    papers_json = json.dumps(papers_for_ai, ensure_ascii=False, indent=2)

    system_msg = (
        "你是学术论文分析助手。用户已通过相关度筛选，你的任务是为每篇论文撰写可读的中文材料。\n"
        "要求：\n"
        "1. 只根据提供的标题与摘要撰写，禁止编造作者、数据、结论或实验结果。\n"
        "2. summary_cn：2～3 句，说明研究问题、方法、主要发现；避免「本文提出了一种…」式套话。\n"
        "3. recommendation_text：像懂行的朋友做深度点评——说明研究价值、对用户的用处、值得关注的证据；"
        "专业但可读，80～150 字，一段完成，不要分点，不要写「推荐理由」四字。\n"
        "4. 首次出现的英文术语附简短中文解释。\n"
        "5. highlights / limitations 各一句，基于摘要，不夸大。\n"
        "6. results 数组必须包含与输入论文数量相同、id 从 0 连续递增的条目，不得遗漏。\n"
        "必须返回 JSON，仅返回 JSON。"
    )
    user_msg = f"""用户兴趣：{interest}
用户核心关注点：{core_concepts}
用户产出意图：{intent_type}（{_intent_type_cn(intent_type)}）

请对以下论文逐一处理，返回 JSON：
{{
  "results": [
    {{
      "id": 0,
      "title_cn": "中文标题",
      "summary_cn": "2～3 句通俗总结",
      "recommendation_text": "深度点评一段",
      "highlights": "一句亮点",
      "limitations": "一句不足"
    }}
  ]
}}

论文列表：
{papers_json}
"""

    content = call_deepseek(
        api_key,
        [{"role": "system", "content": system_msg}, {"role": "user", "content": user_msg}],
        json_mode=True,
        max_tokens=P3_ENRICH_MAX_TOKENS,
    )
    parsed = parse_json_response(content, stage="P3", context=log_context)
    results = _extract_results_list(parsed)

    enrich_map: dict[int, dict] = {}
    for item in results:
        if not isinstance(item, dict):
            continue
        paper_id = item.get("id", item.get("index"))
        if paper_id is None:
            continue
        try:
            enrich_map[int(paper_id)] = item
        except (TypeError, ValueError):
            continue

    if len(enrich_map) < len(batch):
        print(
            f"[P3] {log_context} 返回条数不足: 期望 {len(batch)} 篇, 实际 {len(enrich_map)} 篇"
        )

    enriched: list[dict] = []
    for i, p in enumerate(batch):
        info = enrich_map.get(i, {})
        if info:
            enriched.append(_apply_enrich_info(p, info))
        else:
            failed = dict(p)
            failed["title_cn"] = ENRICH_FAILED_TITLE_CN
            failed["summary_cn"] = ""
            failed["one_liner"] = ""
            failed["recommendation_text"] = ""
            failed["highlights"] = ""
            failed["limitations"] = ""
            enriched.append(failed)
    return enriched


def enrich_papers_batch(
    api_key: str,
    batch: list[dict],
    interest: str,
    *,
    core_concepts: str = "",
    intent_type: str = "exploration",
    batch_index: int = 0,
    total_batches: int = 0,
) -> list[dict]:
    """P3：生成中文材料；解析失败或缺篇时整批重试，仍失败则单篇重试。"""
    if not batch:
        return []

    batch_label = (
        f"batch={batch_index}/{total_batches} size={len(batch)}"
        if batch_index and total_batches
        else f"size={len(batch)}"
    )
    enriched = _enrich_papers_batch_once(
        api_key,
        batch,
        interest,
        core_concepts=core_concepts,
        intent_type=intent_type,
        log_context=batch_label,
    )

    missing = [i for i, p in enumerate(enriched) if not _paper_enrich_ok(p)]
    if not missing:
        return enriched

    if len(batch) > 1:
        print(f"[P3] {batch_label} 有 {len(missing)} 篇不完整，整批重试...")
        retry = _enrich_papers_batch_once(
            api_key,
            batch,
            interest,
            core_concepts=core_concepts,
            intent_type=intent_type,
            log_context=f"{batch_label} retry",
        )
        for i in missing:
            if _paper_enrich_ok(retry[i]):
                enriched[i] = retry[i]
        missing = [i for i, p in enumerate(enriched) if not _paper_enrich_ok(p)]

    for i in missing:
        print(f"[P3] {batch_label} 单篇重试 index={i} title={batch[i].get('title', '')[:60]}")
        single = _enrich_papers_batch_once(
            api_key,
            [batch[i]],
            interest,
            core_concepts=core_concepts,
            intent_type=intent_type,
            log_context=f"{batch_label} single id={i}",
        )
        if _paper_enrich_ok(single[0]):
            enriched[i] = single[0]

    still_missing = [i for i, p in enumerate(enriched) if not _paper_enrich_ok(p)]
    if still_missing:
        print(f"[P3] {batch_label} 仍有 {len(still_missing)} 篇未能生成中文材料")

    return enriched


def compute_scores_and_select(
    enriched_papers: list[dict],
) -> tuple[list[dict], list[dict]]:
    """计算推荐指数，返回 (推荐阅读 Top10, 全部相关文献按推荐指数降序)。"""
    citations = [p.get("citationCount", 0) or 0 for p in enriched_papers]
    min_cite = min(citations) if citations else 0
    max_cite = max(citations) if citations else 1

    for i, p in enumerate(enriched_papers):
        if max_cite > min_cite:
            norm_cite = (citations[i] - min_cite) / (max_cite - min_cite)
        else:
            norm_cite = 1.0

        relevance = float(p.get("relevance_score", 0.5))
        rec_index = (relevance * 0.5 + norm_cite * 0.5) * 10
        p["normalized_citation_score"] = round(norm_cite, 4)
        p["recommendation_index"] = round(rec_index, 2)
        p["score"] = p["recommendation_index"]

    sorted_papers = sorted(
        enriched_papers,
        key=lambda x: x.get("recommendation_index", 0),
        reverse=True,
    )
    recommended = sorted_papers[:MAX_SELECTED_PAPERS]
    return recommended, sorted_papers


_REVIEW_CONCLUSION_PREFIX_RE = re.compile(r"^\d+[\.\、\)]\s*")


def _normalize_review_overview(text: str) -> str:
    text = (text or "").strip()
    if not text:
        return ""
    if not text.startswith("本期研究围绕"):
        text = text.lstrip("。.，,")
        text = f"本期研究围绕{text}"
    return text


def _normalize_review_themes(text: str) -> str:
    text = (text or "").strip()
    if not text:
        return ""
    if not text.startswith("覆盖"):
        text = text.lstrip("。.，,")
        text = f"覆盖{text}"
    return text


def _normalize_review_conclusions(raw: Any) -> list[str]:
    if not isinstance(raw, list):
        return []
    items: list[str] = []
    for item in raw[:5]:
        line = str(item).strip()
        if not line:
            continue
        line = _REVIEW_CONCLUSION_PREFIX_RE.sub("", line)
        if line:
            items.append(line)
    return items


def assemble_review_text(overview: str, themes: str, conclusions: list[str]) -> str:
    """将结构化字段拼成固定三段格式的综述正文。"""
    overview = _normalize_review_overview(overview)
    themes = _normalize_review_themes(themes)

    numbered: list[str] = []
    for idx, line in enumerate(conclusions, 1):
        numbered.append(f"{idx}. {line}")

    sections: list[str] = []
    if overview:
        sections.append(overview)
    if themes:
        sections.append(themes)
    if numbered:
        sections.append("值得关注的结论有：\n" + "\n".join(numbered))

    return "\n\n".join(sections).strip()


def format_review_text(review_text: str) -> str:
    """兼容旧版纯文本综述的后处理（JSON 解析失败时兜底）。"""
    review_text = (review_text or "").strip()
    if not review_text:
        return review_text

    pos = review_text.find("本期研究围绕")
    if pos > 0:
        review_text = review_text[pos:].lstrip()

    if "覆盖" in review_text:
        review_text = re.sub(r"(?<!\n)覆盖", "\n\n覆盖", review_text, count=1)
    if "值得关注的结论有：" in review_text:
        review_text = re.sub(
            r"(?<!\n)值得关注的结论有：",
            "\n\n值得关注的结论有：",
            review_text,
            count=1,
        )

    review_text = re.sub(r"(?<!\n)(\d+\.\s)", r"\n\1", review_text)

    while "\n\n\n" in review_text:
        review_text = review_text.replace("\n\n\n", "\n\n")

    return review_text.strip()


def generate_review(
    api_key: str,
    selected_papers: list[dict],
    *,
    intent_type: str = "exploration",
) -> str:
    """生成中文综述（仅基于推荐阅读 Top10）。"""
    lines = []
    for p in selected_papers:
        lines.append(f"- {p.get('title_cn', p['title'])}：{p.get('one_liner', '')}")
    summary_input = "\n".join(lines)

    intent_hint = ""
    if intent_type == "report_support":
        intent_hint = (
            "用户意图为报告/课题支撑：首段概括可适度点明这些研究对用户问题的证据价值，"
            "但仍须基于下列论文信息，不要编造未出现的结论。"
        )
    elif intent_type == "content_creation":
        intent_hint = (
            "用户意图为内容创作：首段与结论可侧重哪些发现适合科普转述、哪些证据有传播价值。"
        )
    else:
        intent_hint = "用户意图为领域探索：首段与结论侧重研究趋势与主题分布。"

    system_msg = (
        "你是学术综述撰写专家，用流畅、通俗的中文撰写文献速递综述。"
        "只根据提供的论文标题与总结撰写，禁止编造未出现在输入中的数据、作者或结论。"
        "避免「本文提出…」「综上所述」等套话；写给感兴趣但非该领域专家的读者。"
        "你必须只返回一个合法 JSON 对象，不要输出任何其他文字。"
    )
    user_msg = f"""根据以下推荐阅读论文的中文标题与总结，撰写中文文献速递综述。
{intent_hint}

请返回 JSON 对象（仅 JSON），包含以下字段：
- overview: 字符串。以「本期研究围绕」开头，概括本期研究方向（仅这一段，不要写主题分布与结论）。
- themes: 字符串。以「覆盖」开头，描述主要主题分布（仅这一段）。
- conclusions: 字符串数组。1～5 条结论要点，每条为完整句子，不要带序号前缀。

三部分合计 200-400 字，语言流畅自然。

论文信息：
{summary_input}"""

    content = call_deepseek(
        api_key,
        [{"role": "system", "content": system_msg}, {"role": "user", "content": user_msg}],
        json_mode=True,
        temperature=0.3,
        max_tokens=JSON_MAX_TOKENS,
    )
    data = parse_json_response(content, stage="P5-review")
    if isinstance(data, dict):
        overview = safe_get_json_field(data, "overview", default="", expected_type=str) or ""
        themes = safe_get_json_field(data, "themes", default="", expected_type=str) or ""
        conclusions = _normalize_review_conclusions(
            safe_get_json_field(data, "conclusions", default=[], expected_type=list)
        )
        assembled = assemble_review_text(overview, themes, conclusions)
        if assembled:
            return assembled

    return format_review_text(content)


PAPER_SHEET_HEADERS = [
    "推荐指数",
    "相关度得分（P2）",
    "相关度分级",
    "归一化引用得分",
    "英文标题",
    "中文标题",
    "作者",
    "期刊",
    "发表年份",
    "引用次数",
    "英文摘要",
    "一句话总结（中文）",
    "深度点评",
    "亮点",
    "不足",
]
PAPER_SHEET_COL_WIDTHS = [10, 10, 10, 12, 30, 30, 20, 20, 8, 8, 50, 40, 50, 28, 28]


def _write_paper_sheet_rows(
    ws: Any,
    papers: list[dict],
    *,
    bold: Font,
    wrap: Alignment,
    start_row: int = 2,
) -> None:
    for row_idx, p in enumerate(papers, start_row):
        ws.cell(row=row_idx, column=1, value=p.get("recommendation_index", p.get("score", 0)))
        ws.cell(row=row_idx, column=2, value=p.get("relevance_score", ""))
        ws.cell(row=row_idx, column=3, value=p.get("relevance_tier", ""))
        ws.cell(row=row_idx, column=4, value=p.get("normalized_citation_score", ""))
        ws.cell(row=row_idx, column=5, value=p.get("title", ""))
        ws.cell(row=row_idx, column=6, value=p.get("title_cn", ""))
        ws.cell(row=row_idx, column=7, value=p.get("authors", ""))
        ws.cell(row=row_idx, column=8, value=p.get("journal", ""))
        ws.cell(row=row_idx, column=9, value=p.get("year", ""))
        ws.cell(row=row_idx, column=10, value=p.get("citationCount", 0))
        c10 = ws.cell(row=row_idx, column=11, value=p.get("abstract", ""))
        c10.alignment = wrap
        c11 = ws.cell(row=row_idx, column=12, value=p.get("one_liner", ""))
        c11.alignment = wrap
        c12 = ws.cell(row=row_idx, column=13, value=p.get("recommendation_text", ""))
        c12.alignment = wrap
        c13 = ws.cell(row=row_idx, column=14, value=p.get("highlights", ""))
        c13.alignment = wrap
        c14 = ws.cell(row=row_idx, column=15, value=p.get("limitations", ""))
        c14.alignment = wrap


def _init_paper_sheet(ws: Any, *, bold: Font) -> None:
    for col, (h, w) in enumerate(zip(PAPER_SHEET_HEADERS, PAPER_SHEET_COL_WIDTHS), 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = bold
        ws.column_dimensions[get_column_letter(col)].width = w


def build_excel(
    review: str,
    recommended: list[dict],
    all_relevant: list[dict],
) -> bytes:
    """生成三 Sheet：文献速递 / 推荐阅读 / 全部相关论文。"""
    wb = Workbook()
    bold = Font(bold=True)
    wrap = Alignment(wrap_text=True, vertical="top")
    last_col = get_column_letter(len(PAPER_SHEET_HEADERS))

    ws1 = wb.active
    ws1.title = "文献速递"
    ws1.merge_cells(f"A1:{last_col}1")
    cell = ws1["A1"]
    cell.value = review
    cell.alignment = wrap
    ws1.row_dimensions[1].height = 120
    ws1["A3"] = (
        f"以上综述基于推荐阅读的 {len(recommended)} 篇论文；"
        f"共 {len(all_relevant)} 篇相关论文见「全部相关论文」，推荐阅读见「推荐阅读」。"
    )

    ws2 = wb.create_sheet("推荐阅读")
    _init_paper_sheet(ws2, bold=bold)
    if recommended:
        _write_paper_sheet_rows(ws2, recommended, bold=bold, wrap=wrap)
    else:
        ws2["A2"] = "无"

    ws3 = wb.create_sheet("全部相关论文")
    _init_paper_sheet(ws3, bold=bold)
    if all_relevant:
        _write_paper_sheet_rows(ws3, all_relevant, bold=bold, wrap=wrap)
    else:
        ws3["A2"] = "无"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def format_paper_for_api(p: dict) -> dict:
    """将内部论文字典格式化为 API 响应字段。"""
    recommendation_index = p.get("recommendation_index", p.get("score", 0))
    return {
        "title": p.get("title", ""),
        "title_cn": p.get("title_cn", ""),
        "authors": p.get("authors", ""),
        "journal": p.get("journal", ""),
        "year": p.get("year", 0),
        "citation_count": p.get("citationCount", 0),
        "abstract": p.get("abstract", ""),
        "one_liner": p.get("one_liner", p.get("summary_cn", "")),
        "relevance_score": p.get("relevance_score", 0),
        "normalized_citation_score": p.get("normalized_citation_score", 0),
        "recommendation_index": recommendation_index,
        "recommendation_text": p.get("recommendation_text", ""),
        "relevance_tier": p.get("relevance_tier", ""),
        "highlights": p.get("highlights", ""),
        "limitations": p.get("limitations", ""),
        "score": recommendation_index,
    }


def run_pipeline(
    api_key: str,
    keywords: str,
    journal_input: str,
    time_label: str,
    interest: str,
    status_callback: Callable[[int, str], None] | None = None,
    cancel_check: Callable[[], bool] | None = None,
) -> dict[str, Any]:
    """
    主处理流程，返回完整结果字典。
    status_callback: 接收 (step 1-5, message) 用于任务进度更新。
    cancel_check: 返回 True 时抛出 TaskCancelledError 终止流水线。
    """

    def report(step: int, msg: str) -> None:
        print(msg)
        if status_callback:
            status_callback(step, msg)

    def ensure_not_cancelled() -> None:
        if cancel_check and cancel_check():
            raise TaskCancelledError("搜索已取消")

    keywords = keywords or ""
    journal_input = journal_input or ""

    start_date, end_date = get_date_range(time_label)
    start_date_str = start_date.strftime("%Y-%m-%d")
    end_date_str = end_date.strftime("%Y-%m-%d")

    ensure_not_cancelled()
    report(1, " 正在拆解兴趣，拓展搜索词...")
    p1_result = expand_keywords(api_key, keywords, interest, journal_input)
    ensure_not_cancelled()
    journal_names = p1_result["journal_names"]
    simple_terms = p1_result["simple_terms"]
    simple_terms_display = p1_result["simple_terms_display"]
    core_concepts = p1_result["core_concepts"]
    exclusion_terms = p1_result["exclusion_terms"]
    scoring_guide = p1_result["scoring_guide"]
    intent_type = p1_result.get("intent_type", "exploration")

    print("========== P1 返回 ==========")
    print("simple_terms:", p1_result.get("simple_terms"))
    print("exclusion_terms:", p1_result.get("exclusion_terms"))
    print("journal_names:", p1_result.get("journal_names"))

    journal_names_for_fetch = journal_names if journal_input.strip() else []

    def on_fetch_progress(count: int) -> None:
        report(2, f" 正在从 OpenAlex 抓取论文...（已获取 {count} 篇）")

    report(2, " 正在从 OpenAlex 抓取论文...")
    query = simple_terms
    journals_param = journal_names_for_fetch
    date_from = start_date_str
    date_to = end_date_str
    print("========== OpenAlex 查询串 ==========")
    print(f"查询: {query}")
    print(f"期刊过滤: {journals_param if journals_param else '无'}")
    print(f"时间范围: {date_from} 至 {date_to}")
    print("================================")
    raw_papers = fetch_papers_openalex(
        simple_terms,
        start_date_str,
        end_date_str,
        journal_names_for_fetch,
        progress_callback=on_fetch_progress,
        cancel_check=cancel_check,
    )
    ensure_not_cancelled()
    if exclusion_terms:
        before = len(raw_papers)
        raw_papers = filter_papers_by_exclusion(raw_papers, exclusion_terms)
        print(f" 排除词过滤：{before} -> {len(raw_papers)} 篇")

    if not raw_papers:
        raise ValueError("未找到符合日期或期刊条件的论文，请调整关键词或时间范围后重试。")

    fetched_count = len(raw_papers)

    report(3, " AI 正在评估相关度...")
    candidate_papers: list[dict] = []
    batch_size_filter = 10
    total_batches = (len(raw_papers) + batch_size_filter - 1) // batch_size_filter
    for b in range(total_batches):
        ensure_not_cancelled()
        batch = raw_papers[b * batch_size_filter : (b + 1) * batch_size_filter]
        filtered = filter_papers_batch(
            api_key,
            batch,
            interest,
            b + 1,
            core_concepts=core_concepts,
            scoring_guide=scoring_guide,
            intent_type=intent_type,
        )
        candidate_papers.extend(filtered)
        report(
            3,
            f" AI 正在评估相关度...（第 {b + 1}/{total_batches} 批，已保留 {len(candidate_papers)} 篇）",
        )
        time.sleep(0.5)

    if not candidate_papers:
        raise ValueError("未找到高度相关论文，请调整兴趣描述后重试。")

    relevant_count = len(candidate_papers)

    report(4, " 正在生成中文摘要与深度点评...")
    enriched_all: list[dict] = []
    batch_size_enrich = BATCH_SIZE_ENRICH
    total_enrich = (len(candidate_papers) + batch_size_enrich - 1) // batch_size_enrich
    for b in range(total_enrich):
        ensure_not_cancelled()
        batch = candidate_papers[b * batch_size_enrich : (b + 1) * batch_size_enrich]
        enriched = enrich_papers_batch(
            api_key,
            batch,
            interest,
            core_concepts=core_concepts,
            intent_type=intent_type,
            batch_index=b + 1,
            total_batches=total_enrich,
        )
        enriched_all.extend(enriched)
        report(
            4,
            f" 正在生成中文摘要与深度点评...（第 {b + 1}/{total_enrich} 批）",
        )
        time.sleep(0.5)

    ensure_not_cancelled()
    recommended_papers, all_relevant_papers = compute_scores_and_select(enriched_all)
    recommended_count = len(recommended_papers)
    selection_ratio = (
        round(recommended_count / relevant_count, 4) if relevant_count > 0 else 0.0
    )

    ensure_not_cancelled()
    report(5, " 正在生成综述...")
    review = generate_review(
        api_key, recommended_papers, intent_type=intent_type
    )

    ensure_not_cancelled()
    report(5, " 正在生成 Excel...")
    excel_bytes = build_excel(review, recommended_papers, all_relevant_papers)

    journals_used: list[str] = []
    if journal_input.strip():
        journals_used = [j.strip() for j in journal_input.split(",") if j.strip()]
    elif journal_names:
        journals_used = journal_names

    return {
        "summary": review,
        "selected_papers": [format_paper_for_api(p) for p in recommended_papers],
        "search_stats": {
            "fetched_count": fetched_count,
            "relevant_count": relevant_count,
            "recommended_count": recommended_count,
            "selection_ratio": selection_ratio,
        },
        "simple_terms": simple_terms_display,
        "journals_used": journals_used,
        "time_range": time_label,
        "interest": interest,
        "excel_bytes": excel_bytes,
    }


# ---------------------------------------------------------------------------
# FastAPI 应用
# ---------------------------------------------------------------------------

app = FastAPI(title="LOOKLOOK API", description="文献检索与智能速递后端")

TEMPLATES_DIR = Path(__file__).parent / "templates"


@app.on_event("startup")
def on_startup() -> None:
    if _use_postgres():
        init_db_pool()
    init_db()


@app.on_event("shutdown")
def on_shutdown() -> None:
    close_db_pool()


class SearchRequest(BaseModel):
    api_key: str = Field(default="", description="DeepSeek API Key（免 Key 模式可省略）")
    interest: str = Field(..., description="个人兴趣描述")
    keywords: str = Field(default="", description="可选关键词")
    journal_input: str = Field(default="", description="可选期刊名")
    time_range: str = Field(default="三年内", description="时间范围")
    session_id: str = Field(default="", description="匿名用户会话标识")


class UsageRemainingResponse(BaseModel):
    limit: int
    used: int
    remaining: int
    logged_in: bool


class TaskStartResponse(BaseModel):
    task_id: str


class TaskStatusResponse(BaseModel):
    step: int
    message: str
    done: bool
    cancelled: bool = False
    result: dict[str, Any] | None = None
    error: str | None = None


class RegisterRequest(BaseModel):
    email: str = Field(..., description="用户邮箱")
    username: str = Field(..., description="用户名")
    password: str = Field(..., description="登录密码")


class LoginRequest(BaseModel):
    email: str = Field(..., description="用户邮箱")
    password: str = Field(..., description="登录密码")


class RegisterResponse(BaseModel):
    message: str
    user_id: int


class LoginResponse(BaseModel):
    token: str
    user_id: int
    username: str


class MeResponse(BaseModel):
    user_id: int
    email: str
    username: str


class HistoryItemResponse(BaseModel):
    id: int
    interest_desc: str | None = None
    keywords: str | None = None
    simple_terms: str | None = None
    time_range: str | None = None
    journals: str | None = None
    summary: str | None = None
    created_at: str


def _update_task(
    task_id: str,
    *,
    step: int | None = None,
    message: str | None = None,
    done: bool | None = None,
    cancelled: bool | None = None,
    result: dict[str, Any] | None = None,
    error: str | None = None,
) -> None:
    with task_lock:
        if task_id not in task_status:
            return
        if step is not None:
            task_status[task_id]["step"] = step
        if message is not None:
            task_status[task_id]["message"] = message
        if done is not None:
            task_status[task_id]["done"] = done
        if cancelled is not None:
            task_status[task_id]["cancelled"] = cancelled
        if result is not None:
            task_status[task_id]["result"] = result
        if error is not None:
            task_status[task_id]["error"] = error


def _is_task_cancelled(task_id: str) -> bool:
    with task_lock:
        task = task_status.get(task_id)
        return bool(task and task.get("cancelled"))


def request_cancel_task(task_id: str) -> bool:
    """标记任务为已取消；任务尚未结束时返回 True。"""
    with task_lock:
        task = task_status.get(task_id)
        if not task or task.get("done"):
            return False
        task["cancelled"] = True
        return True


def _finalize_cancelled_task(task_id: str) -> None:
    _update_task(
        task_id,
        done=True,
        cancelled=True,
        message="搜索已取消",
        error=None,
        result=None,
    )


def _run_search_task(task_id: str, req: SearchRequest, user_id: int | None = None) -> None:
    """后台线程执行流水线，更新 task_status。"""

    def on_status(step: int, message: str) -> None:
        if _is_task_cancelled(task_id):
            raise TaskCancelledError("搜索已取消")
        _update_task(task_id, step=step, message=message)

    def cancel_check() -> bool:
        return _is_task_cancelled(task_id)

    try:
        _update_task(task_id, step=0, message="正在准备...")
        if _is_task_cancelled(task_id):
            _finalize_cancelled_task(task_id)
            return
        result = run_pipeline(
            api_key=req.api_key.strip(),
            keywords=req.keywords.strip() if req.keywords else "",
            journal_input=req.journal_input.strip() if req.journal_input else "",
            time_label=req.time_range,
            interest=req.interest.strip(),
            status_callback=on_status,
            cancel_check=cancel_check,
        )
        if _is_task_cancelled(task_id):
            _finalize_cancelled_task(task_id)
            return
        excel_bytes = result.pop("excel_bytes")
        result["excel_base64"] = base64.b64encode(excel_bytes).decode("utf-8")
        if user_id is not None:
            save_search_history(user_id, req, result)
        _update_task(
            task_id,
            step=5,
            message=" 分析完成！",
            done=True,
            result=result,
        )
    except TaskCancelledError:
        _finalize_cancelled_task(task_id)
    except ValueError as e:
        if _is_task_cancelled(task_id):
            _finalize_cancelled_task(task_id)
            return
        _update_task(task_id, done=True, error=str(e), message=str(e))
    except Exception as e:
        if _is_task_cancelled(task_id):
            _finalize_cancelled_task(task_id)
            return
        _update_task(
            task_id,
            done=True,
            error=f"流水线执行失败: {e}",
            message=f"流水线执行失败: {e}",
        )


@app.get("/")
def index_page():
    """返回前端单页应用。"""
    html_path = TEMPLATES_DIR / "index.html"
    if not html_path.is_file():
        raise HTTPException(status_code=404, detail="前端页面未找到")
    return FileResponse(html_path, media_type="text/html; charset=utf-8")


@app.get("/health")
def health():
    return {"status": "healthy"}


@app.post("/api/register", response_model=RegisterResponse)
def api_register(req: RegisterRequest):
    email = _normalize_email(req.email)
    username = req.username.strip()
    password = req.password

    if not _is_valid_email(email):
        raise HTTPException(status_code=400, detail="邮箱格式无效")
    if not username:
        raise HTTPException(status_code=400, detail="用户名不能为空")
    if not password:
        raise HTTPException(status_code=400, detail="密码不能为空")

    if _get_user_by_email(email) is not None:
        raise HTTPException(status_code=400, detail="该邮箱已注册")

    password_hash = hash_password(password)
    user_id: int | None = None
    try:
        with get_db_connection() as conn:
            if _use_postgres():
                row = _db_execute(
                    conn,
                    "INSERT INTO users (email, username, password_hash) VALUES (?, ?, ?) RETURNING id",
                    (email, username, password_hash),
                ).fetchone()
                user_id = int(row["id"]) if row else None
            else:
                cursor = _db_execute(
                    conn,
                    "INSERT INTO users (email, username, password_hash) VALUES (?, ?, ?)",
                    (email, username, password_hash),
                )
                user_id = cursor.lastrowid
            conn.commit()
    except Exception as e:
        if _is_unique_violation(e):
            raise HTTPException(status_code=400, detail="该邮箱已注册") from None
        raise

    if user_id is None:
        raise HTTPException(status_code=500, detail="注册失败，请稍后重试")

    return RegisterResponse(message="注册成功", user_id=user_id)


@app.get("/api/me", response_model=MeResponse)
def api_me(authorization: str | None = Header(default=None)):
    user_id = _require_user_id_from_header(authorization)
    user = _get_user_by_id(user_id)
    if user is None:
        raise HTTPException(status_code=401, detail="用户不存在")
    return MeResponse(
        user_id=user["id"],
        email=user["email"],
        username=user["username"],
    )


@app.get("/api/user/history", response_model=list[HistoryItemResponse])
def api_user_history(authorization: str | None = Header(default=None)):
    """返回当前登录用户的检索记录，按时间倒序。"""
    user_id = _require_user_id_from_header(authorization)
    with get_db_connection() as conn:
        rows = _db_execute(
            conn,
            """
            SELECT id, interest_desc, keywords, simple_terms, time_range,
                   journals, summary, created_at
            FROM search_history
            WHERE user_id = ?
            ORDER BY created_at DESC
            """,
            (user_id,),
        ).fetchall()
    return [
        HistoryItemResponse(
            id=row["id"],
            interest_desc=row["interest_desc"],
            keywords=row["keywords"],
            simple_terms=row["simple_terms"],
            time_range=row["time_range"],
            journals=row["journals"],
            summary=row["summary"],
            created_at=str(row["created_at"]),
        )
        for row in rows
    ]


@app.post("/api/login", response_model=LoginResponse)
def api_login(req: LoginRequest):
    email = _normalize_email(req.email)
    password = req.password

    if not email or not password:
        raise HTTPException(status_code=400, detail="邮箱和密码不能为空")

    user = _get_user_by_email(email)
    if user is None or not verify_password(password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="邮箱或密码错误")

    token = create_access_token(user["id"], user["email"])
    return LoginResponse(
        token=token,
        user_id=user["id"],
        username=user["username"],
    )


@app.get("/api/mode")
def api_mode():
    """返回当前是否为免 Key 模式。"""
    return {"free_mode": FREE_MODE}


def _resolve_api_key(client_api_key: str) -> str:
    """免 Key 模式使用服务端默认 Key，否则要求客户端提供。"""
    if FREE_MODE:
        key = DEFAULT_DEEPSEEK_KEY.strip()
        if not key:
            raise HTTPException(
                status_code=500,
                detail="服务器未配置默认 DeepSeek API Key，请联系管理员",
            )
        return key
    if not client_api_key or not client_api_key.strip():
        raise HTTPException(status_code=400, detail="请提供有效的 DeepSeek API Key")
    return client_api_key.strip()


@app.get("/api/usage/remaining", response_model=UsageRemainingResponse)
def api_usage_remaining(
    session_id: str = "",
    authorization: str | None = Header(default=None),
):
    """查询当前用户今日剩余搜索次数（不扣减配额）。"""
    user_id = _optional_user_id_from_header(authorization)
    snapshot = build_usage_snapshot(
        user_id=user_id,
        identifier=session_id,
    )
    return UsageRemainingResponse(**snapshot)


@app.post("/api/search", response_model=TaskStartResponse)
def api_search(
    req: SearchRequest,
    request: Request,
    authorization: str | None = Header(default=None),
):
    """创建异步检索任务，立即返回 task_id。"""
    if not req.interest or not req.interest.strip():
        raise HTTPException(status_code=400, detail="请提供个人兴趣描述")
    if req.time_range not in TIME_RANGE_OPTIONS:
        raise HTTPException(
            status_code=400,
            detail=f"无效的时间范围，可选: {', '.join(TIME_RANGE_OPTIONS)}",
        )

    user_id = _optional_user_id_from_header(authorization)
    if user_id is not None:
        if not consume_user_usage(user_id):
            raise HTTPException(
                status_code=429,
                detail="您今日的搜索次数已用完（20次/天），请明天再来。",
            )
    else:
        client_host = request.client.host if request.client else None
        identifier = _resolve_anonymous_identifier(req.session_id, client_host)
        if not consume_anonymous_usage(identifier):
            raise HTTPException(
                status_code=429,
                detail="免费用户每日可搜索3次，您今日的次数已用完。登录后每天可搜索20次。",
            )

    effective_api_key = _resolve_api_key(req.api_key)
    # 已登录用户：流水线成功后在后台线程写入 search_history

    task_id = str(uuid.uuid4())
    with task_lock:
        task_status[task_id] = {
            "step": 0,
            "message": "任务已创建，等待执行...",
            "done": False,
            "cancelled": False,
            "result": None,
            "error": None,
        }

    req.api_key = effective_api_key
    thread = threading.Thread(
        target=_run_search_task,
        args=(task_id, req, user_id),
        daemon=True,
    )
    thread.start()
    return TaskStartResponse(task_id=task_id)


@app.post("/api/search/cancel")
def api_search_cancel(task_id: str = Query(..., description="要取消的任务 ID")):
    """取消进行中的搜索任务（关闭弹窗或刷新页面时调用）。"""
    if not request_cancel_task(task_id):
        with task_lock:
            if task_id not in task_status:
                raise HTTPException(status_code=404, detail="任务不存在或已过期")
        return {"ok": True, "message": "任务已结束"}
    return {"ok": True, "message": "已请求取消"}


@app.get("/api/search/status", response_model=TaskStatusResponse)
def api_search_status(task_id: str):
    """查询异步任务进度与结果。"""
    with task_lock:
        if task_id not in task_status:
            raise HTTPException(status_code=404, detail="任务不存在或已过期")
        task = task_status[task_id].copy()

    return TaskStatusResponse(
        step=task.get("step", 0),
        message=task.get("message", ""),
        done=task.get("done", False),
        cancelled=bool(task.get("cancelled")),
        result=task.get("result"),
        error=task.get("error"),
    )


if __name__ == "__main__":
    uvicorn.run(app, host="0.0.0.0", port=8000)
